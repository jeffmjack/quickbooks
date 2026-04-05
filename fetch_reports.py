"""
Fetch QBO reports and save as Excel files in private-docs/.
"""

import os
import requests
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

BASE_URL = "https://quickbooks.api.intuit.com"


def get_access_token():
    resp = requests.post(
        "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer",
        auth=(os.getenv("QBO_CLIENT_ID"), os.getenv("QBO_CLIENT_SECRET")),
        data={"grant_type": "refresh_token", "refresh_token": os.getenv("QBO_REFRESH_TOKEN")},
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def get_headers(token):
    return {"Authorization": f"Bearer {token}", "Accept": "application/json"}


def fetch_report(headers, realm_id, report_name, params):
    resp = requests.get(
        f"{BASE_URL}/v3/company/{realm_id}/reports/{report_name}",
        headers=headers,
        params=params,
    )
    resp.raise_for_status()
    return resp.json()


def fetch_query(headers, realm_id, query):
    resp = requests.get(
        f"{BASE_URL}/v3/company/{realm_id}/query",
        headers=headers,
        params={"query": query},
    )
    resp.raise_for_status()
    return resp.json()


def parse_report_to_df(data):
    """Parse a QBO report JSON into a flat DataFrame."""
    cols = [c.get("ColTitle", "") for c in data["Columns"]["Column"]]
    rows = []

    def walk(row_list, indent=0):
        for row in row_list:
            row_type = row.get("type")
            if row_type == "Section":
                header = row.get("Header", {}).get("ColData", [])
                if header:
                    values = [("  " * indent) + header[0].get("value", "")] + \
                              [d.get("value", "") for d in header[1:]]
                    rows.append(values)
                if "Rows" in row:
                    walk(row["Rows"].get("Row", []), indent + 1)
                summary = row.get("Summary", {}).get("ColData", [])
                if summary:
                    values = [("  " * indent) + summary[0].get("value", "")] + \
                              [d.get("value", "") for d in summary[1:]]
                    rows.append(values)
            elif row_type == "Data":
                col_data = row.get("ColData", [])
                values = [("  " * indent) + col_data[0].get("value", "")] + \
                          [d.get("value", "") for d in col_data[1:]]
                rows.append(values)

    walk(data["Rows"]["Row"])

    # Pad rows to match column count
    ncols = len(cols)
    rows = [r + [""] * (ncols - len(r)) if len(r) < ncols else r for r in rows]
    return pd.DataFrame(rows, columns=cols)


def style_sheet(ws):
    """Apply basic formatting to a worksheet."""
    # Header row
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")
    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)


def save_df_to_excel(df, path, sheet_name="Report"):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        style_sheet(writer.sheets[sheet_name])
    print(f"  Saved: {path}")


def main():
    os.makedirs("private-docs", exist_ok=True)
    token = get_access_token()
    headers = get_headers(token)
    realm_id = os.getenv("QBO_REALM_ID")

    # 1. P&L by month: 1/1/2025 - 2/28/2026
    print("Fetching P&L by month...")
    data = fetch_report(headers, realm_id, "ProfitAndLoss", {
        "start_date": "2025-01-01",
        "end_date": "2026-02-28",
        "accounting_method": "Cash",
        "summarize_column_by": "Month",
    })
    save_df_to_excel(parse_report_to_df(data),
                     "private-docs/ProfitAndLoss_by_month_2025-2026.xlsx", "P&L by Month")

    # 2. Balance Sheet by month: 1/1/2025 - 2/28/2026
    print("Fetching Balance Sheet by month...")
    data = fetch_report(headers, realm_id, "BalanceSheet", {
        "start_date": "2025-01-01",
        "end_date": "2026-02-28",
        "summarize_column_by": "Month",
    })
    save_df_to_excel(parse_report_to_df(data),
                     "private-docs/BalanceSheet_by_month_2025-2026.xlsx", "Balance Sheet by Month")

    # 3. General Ledger: 1/1/2026 - 2/28/2026 (fetched month by month to avoid timeout)
    print("Fetching General Ledger (Jan 2026)...")
    gl_jan = fetch_report(headers, realm_id, "GeneralLedger", {
        "start_date": "2026-01-01", "end_date": "2026-01-31",
    })
    print("Fetching General Ledger (Feb 2026)...")
    gl_feb = fetch_report(headers, realm_id, "GeneralLedger", {
        "start_date": "2026-02-01", "end_date": "2026-02-28",
    })
    df_gl = pd.concat([parse_report_to_df(gl_jan), parse_report_to_df(gl_feb)], ignore_index=True)
    save_df_to_excel(df_gl, "private-docs/GeneralLedger_Jan-Feb_2026.xlsx", "General Ledger")

    # 4. Sales by Customer Detail: 1/1/2026 - 2/28/2026
    print("Fetching Sales by Customer...")
    data = fetch_report(headers, realm_id, "CustomerSales", {
        "start_date": "2026-01-01",
        "end_date": "2026-02-28",
        "accounting_method": "Cash",
    })
    save_df_to_excel(parse_report_to_df(data),
                     "private-docs/SalesByCustomer_Jan-Feb_2026.xlsx", "Sales by Customer")

    # 5. AR Aging as of 2/28/2026
    print("Fetching AR Aging...")
    data = fetch_report(headers, realm_id, "AgedReceivableDetail", {
        "report_date": "2026-02-28",
        "aging_period": "30",
        "num_periods": "4",
    })
    save_df_to_excel(parse_report_to_df(data),
                     "private-docs/AR_Aging_2026-02-28.xlsx", "AR Aging")

    # 6. Bill Payment List: 1/1/2026 - 2/28/2026
    print("Fetching Bill Payments...")
    result = fetch_query(headers, realm_id,
        "SELECT * FROM BillPayment WHERE TxnDate >= '2026-01-01' AND TxnDate <= '2026-02-28' MAXRESULTS 1000"
    )
    bill_payments = result.get("QueryResponse", {}).get("BillPayment", [])
    if bill_payments:
        rows = []
        for bp in bill_payments:
            for line in bp.get("Line", []):
                for linked in line.get("LinkedTxn", []):
                    rows.append({
                        "Date": bp.get("TxnDate"),
                        "Vendor": bp.get("VendorRef", {}).get("name", ""),
                        "Payment Method": bp.get("PayType", ""),
                        "Amount": bp.get("TotalAmt", ""),
                        "Bill Date": "",
                        "Bill Amount": line.get("Amount", ""),
                        "Linked Txn ID": linked.get("TxnId", ""),
                        "Memo": bp.get("PrivateNote", ""),
                    })
        df = pd.DataFrame(rows)
    else:
        df = pd.DataFrame(columns=["Date", "Vendor", "Payment Method", "Amount", "Memo"])
    save_df_to_excel(df, "private-docs/BillPayments_Jan-Feb_2026.xlsx", "Bill Payments")

    # 7. AP Aging as of 2/28/2026
    print("Fetching AP Aging...")
    data = fetch_report(headers, realm_id, "AgedPayableDetail", {
        "report_date": "2026-02-28",
        "aging_period": "30",
        "num_periods": "4",
    })
    save_df_to_excel(parse_report_to_df(data),
                     "private-docs/AP_Aging_2026-02-28.xlsx", "AP Aging")

    print("\nAll reports saved to private-docs/")


if __name__ == "__main__":
    main()
